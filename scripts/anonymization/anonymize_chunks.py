"""
Anonymize raw clinical chunks before indexing.

Input:
  data/chunks/chunks.raw.jsonl

Output:
  data/chunks/chunks.anonymized.jsonl
  data/chunks/anonymization_report.json
  data/private/anonymization_mapping.xlsx (preferred) or .csv fallback
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

SCHEMA_VERSION = "clinical_chunk_anonymized_v1"

REQUIRED_CHUNK_FIELDS = {
    "chunk_id",
    "doc_id",
    "chunk_type",
    "text_for_embedding",
    "text_for_keyword",
    "metadata",
    "provenance",
    "routing",
}

FORBIDDEN_METADATA_FIELDS = {
    "patient_id",
    "patient_id_raw",
    "ip_patient",
    "patient_birth_date",
    "patient_birth_date_raw",
    "patient_address",
    "prescriber",
    "validated_by",
    "edited_by",
    "printed_by",
}

TOKEN_TARGETS = {
    "patient": ("patient_id", "patient_id_raw", "ip_patient"),
    "sample": ("sample_id", "sample_id_raw"),
    "report": ("report_id", "report_id_raw"),
}

TOKEN_PREFIX = {
    "patient": "PAT",
    "sample": "SAMPLE",
    "report": "REPORT",
}

MAPPING_COLUMNS = [
    "field_name",
    "entity_type",
    "original_value",
    "method",
    "anonymized_value",
    "reversible",
    "created_at",
    "source_doc_ids",
    "notes",
]


@dataclass
class ReplaceSpec:
    original: str
    replacement: str
    is_explicit_id: bool = False
    entity_type: str = ""
    field_name: str = ""


class MappingStore:
    def __init__(self) -> None:
        self.rows: List[Dict[str, str]] = []
        self.by_key: Dict[Tuple[str, str], Dict[str, str]] = {}
        self.next_index: Dict[str, int] = defaultdict(lambda: 1)

    def _merge_source_doc(self, row: Dict[str, str], doc_id: str) -> None:
        current = set(x.strip() for x in (row.get("source_doc_ids") or "").split(",") if x.strip())
        if doc_id:
            current.add(doc_id)
        row["source_doc_ids"] = ", ".join(sorted(current))

    def _merge_field_name(self, row: Dict[str, str], field_name: str) -> None:
        existing = set(x.strip() for x in (row.get("field_name") or "").split(",") if x.strip())
        if field_name:
            existing.add(field_name)
        row["field_name"] = ", ".join(sorted(existing))

    def _recompute_next_indexes(self) -> None:
        for entity in TOKEN_PREFIX:
            prefix = TOKEN_PREFIX[entity] + "_"
            max_seen = 0
            for row in self.rows:
                if row.get("entity_type") != entity:
                    continue
                token = row.get("anonymized_value") or ""
                if token.startswith(prefix):
                    tail = token[len(prefix):]
                    if tail.isdigit():
                        max_seen = max(max_seen, int(tail))
            self.next_index[entity] = max_seen + 1

    def load_from_rows(self, rows: Iterable[Dict[str, Any]]) -> None:
        self.rows = []
        self.by_key = {}

        for raw in rows:
            row = {col: str(raw.get(col, "") or "") for col in MAPPING_COLUMNS}
            key = (row["entity_type"], row["original_value"])
            if not key[0] or not key[1]:
                continue

            existing = self.by_key.get(key)
            if existing:
                self._merge_source_doc(existing, row.get("source_doc_ids", ""))
                self._merge_field_name(existing, row.get("field_name", ""))
                continue

            self.rows.append(row)
            self.by_key[key] = row

        self._recompute_next_indexes()

    def get_or_create_token(
        self,
        *,
        entity_type: str,
        field_name: str,
        original_value: str,
        doc_id: str,
        notes: str = "",
    ) -> Tuple[str, bool]:
        key = (entity_type, original_value)
        row = self.by_key.get(key)

        if row:
            self._merge_source_doc(row, doc_id)
            self._merge_field_name(row, field_name)
            return row["anonymized_value"], True

        index = self.next_index[entity_type]
        self.next_index[entity_type] += 1
        token = f"{TOKEN_PREFIX[entity_type]}_{index:06d}"

        created_at = datetime.now(timezone.utc).isoformat()
        new_row = {
            "field_name": field_name,
            "entity_type": entity_type,
            "original_value": original_value,
            "method": "token_sequence",
            "anonymized_value": token,
            "reversible": "yes",
            "created_at": created_at,
            "source_doc_ids": doc_id,
            "notes": notes,
        }
        self.rows.append(new_row)
        self.by_key[key] = new_row
        return token, False


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, (int, float)):
        return str(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def load_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            text = line.strip()
            if not text:
                continue
            try:
                rows.append(json.loads(text))
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSON at line {line_no}: {exc}") from exc
    return rows


def write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def write_json(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)


def compute_content_hash(chunk: Dict[str, Any]) -> str:
    """
    Deterministic content hash for the final anonymized chunk payload.
    The hash must reflect anonymized content, so content_hash itself is excluded.
    """
    payload = dict(chunk or {})
    payload.pop("content_hash", None)
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def try_load_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception:
        return None


def read_mapping(mapping_xlsx: Path) -> Tuple[MappingStore, str, List[str]]:
    warnings: List[str] = []
    store = MappingStore()

    openpyxl = try_load_openpyxl()
    mapping_csv = mapping_xlsx.with_suffix(".csv")

    if mapping_xlsx.exists() and openpyxl is not None:
        wb = openpyxl.load_workbook(mapping_xlsx)
        ws = wb.active
        rows: List[Dict[str, Any]] = []
        header = [clean(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        col_index = {name: idx for idx, name in enumerate(header)}

        for vals in ws.iter_rows(min_row=2, values_only=True):
            if vals is None:
                continue
            row = {}
            for col in MAPPING_COLUMNS:
                idx = col_index.get(col)
                row[col] = "" if idx is None else clean(vals[idx])
            if not row.get("entity_type") or not row.get("original_value"):
                continue
            rows.append(row)

        store.load_from_rows(rows)
        return store, str(mapping_xlsx), warnings

    if mapping_xlsx.exists() and openpyxl is None:
        warnings.append(
            "openpyxl is not installed, cannot read existing XLSX mapping. Falling back to CSV if available."
        )

    if mapping_csv.exists():
        with mapping_csv.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            rows = []
            for row in reader:
                rows.append({col: clean(row.get(col)) for col in MAPPING_COLUMNS})
        store.load_from_rows(rows)
        return store, str(mapping_csv), warnings

    if openpyxl is None:
        warnings.append(
            "openpyxl not available. Mapping will be written to CSV fallback instead of XLSX."
        )
    return store, str(mapping_xlsx), warnings


def save_mapping(mapping_xlsx: Path, store: MappingStore) -> Tuple[str, List[str]]:
    warnings: List[str] = []
    openpyxl = try_load_openpyxl()
    mapping_xlsx.parent.mkdir(parents=True, exist_ok=True)

    if openpyxl is not None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "anonymization_mapping"
        ws.append(MAPPING_COLUMNS)
        for row in sorted(store.rows, key=lambda r: (r.get("entity_type", ""), r.get("anonymized_value", ""))):
            ws.append([row.get(col, "") for col in MAPPING_COLUMNS])
        wb.save(mapping_xlsx)
        return str(mapping_xlsx), warnings

    mapping_csv = mapping_xlsx.with_suffix(".csv")
    warnings.append(
        f"openpyxl not available. Mapping saved to CSV fallback: {mapping_csv}"
    )
    with mapping_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=MAPPING_COLUMNS)
        writer.writeheader()
        for row in sorted(store.rows, key=lambda r: (r.get("entity_type", ""), r.get("anonymized_value", ""))):
            writer.writerow({col: row.get(col, "") for col in MAPPING_COLUMNS})
    return str(mapping_csv), warnings


TOKEN_RE = re.compile(r"\b(?:PAT|SAMPLE|REPORT)_\d{6}\b")
SHORT_NUMERIC_BLOCKLIST = {"0", "1", "2", "53"}


def is_safe_global_replacement(value: str) -> bool:
    v = clean(value)
    if not v:
        return False
    if v in SHORT_NUMERIC_BLOCKLIST:
        return False
    if re.fullmatch(r"\d+", v) and len(v) < 6:
        return False
    return True


def _mask_existing_tokens(text: str) -> Tuple[str, Dict[str, str]]:
    placeholders: Dict[str, str] = {}

    def _repl(m: re.Match[str]) -> str:
        key = f"__ANON_TOKEN_{len(placeholders)}__"
        placeholders[key] = m.group(0)
        return key

    return TOKEN_RE.sub(_repl, text), placeholders


def _unmask_existing_tokens(text: str, placeholders: Dict[str, str]) -> str:
    restored = text
    for key, value in placeholders.items():
        restored = restored.replace(key, value)
    return restored


def replace_short_id_with_context(text: str, field_labels: Sequence[str], original_value: str, token: str) -> Tuple[str, int]:
    if not text:
        return text, 0
    value = clean(original_value)
    if not value:
        return text, 0

    labels = [lbl for lbl in field_labels if clean(lbl)]
    if not labels:
        return text, 0

    label_pattern = "|".join(re.escape(lbl) for lbl in labels)
    pattern = re.compile(rf"(?i)\b({label_pattern})\b(\s*(?::|=|-)?\s*){re.escape(value)}(?=\b|$)")

    return pattern.subn(rf"\1\2{token}", text)


def _context_labels_for(spec: ReplaceSpec) -> List[str]:
    if spec.entity_type == "patient":
        return [
            "IP Patient",
            "Patient ID",
            "patient_id",
            "patient_id_raw",
            "ip_patient",
            "N° patient",
            "Identifiant patient",
        ]
    if spec.entity_type == "sample":
        return ["sample_id", "sample id", "id échantillon", "identifiant echantillon"]
    if spec.entity_type == "report":
        return ["report_id", "report id", "id rapport", "identifiant rapport"]
    return []


def anonymize_text_safely(text: str, replacements: Sequence[ReplaceSpec], counters: Dict[str, int]) -> Tuple[str, int]:
    if not text:
        return text, 0

    masked_text, placeholders = _mask_existing_tokens(text)
    ordered = sorted(
        [r for r in replacements if clean(r.original)],
        key=lambda s: len(clean(s.original)),
        reverse=True,
    )

    current = masked_text
    total = 0
    for spec in ordered:
        original = clean(spec.original)
        if not original:
            continue

        if is_safe_global_replacement(original):
            current, count = re.subn(re.escape(original), spec.replacement, current)
            total += count
            continue

        if re.fullmatch(r"\d+", original) and len(original) < 6:
            counters["skipped_short_numeric_replacements_count"] += 1
            labels = _context_labels_for(spec)
            current, count = replace_short_id_with_context(current, labels, original, spec.replacement)
            total += count

    return _unmask_existing_tokens(current, placeholders), total


ADMIN_LABEL_PATTERNS = [
    r"\bEdit(?:e|é|e\u0301)(?:\(e\)|e)?\s+par\b",
    r"\bValid(?:e|é|e\u0301)(?:\(e\)|e)?\s+par\b",
    r"\bImprim(?:e|é|e\u0301)(?:e)?\s+par\b",
    r"\bPrescripteur\b",
    r"\bDr\.",
    r"\bPr\.",
    r"\bM(?:e|é|e\u0301)decin\b",
    r"\bTechnicien\b",
    r"\bBiologiste\b",
]

BLOCKING_ADMIN_PATTERNS = [
    r"\bEdit(?:e|é|e\u0301)(?:\(e\)|e)?\s+par\b",
    r"\bValid(?:e|é|e\u0301)(?:\(e\)|e)?\s+par\b",
    r"\bImprim(?:e|é|e\u0301)(?:e)?\s+par\b",
    r"\bPrescripteur\b",
    r"\bDr\.",
]


def sanitize_administrative_text(text: str) -> str:
    if not text:
        return ""
    out = text
    for pattern in ADMIN_LABEL_PATTERNS:
        out = re.sub(pattern, " ", out, flags=re.IGNORECASE)
    out = re.sub(r"\s+([:;,\.\)])", r"\1", out)
    out = re.sub(r"([(\[])\s+", r"\1", out)
    out = re.sub(r"\s{2,}", " ", out).strip()
    return out


def maybe_sensitive_website(value: str, sensitive_values: Sequence[str]) -> bool:
    v = clean(value)
    if not v:
        return False
    for s in sensitive_values:
        sv = clean(s)
        if sv and sv in v:
            return True
    return False


def build_replacement_specs(
    metadata: Dict[str, Any],
    tokens: Dict[str, str],
) -> Tuple[List[ReplaceSpec], List[str]]:
    specs: List[ReplaceSpec] = []
    originals: List[str] = []

    def add(
        value: Any,
        replacement: str,
        explicit_id: bool = False,
        entity_type: str = "",
        field_name: str = "",
    ) -> None:
        val = clean(value)
        if not val:
            return
        specs.append(
            ReplaceSpec(
                original=val,
                replacement=replacement,
                is_explicit_id=explicit_id,
                entity_type=entity_type,
                field_name=field_name,
            )
        )
        originals.append(val)

    for field in TOKEN_TARGETS["patient"]:
        add(metadata.get(field), tokens.get("patient_token", ""), explicit_id=True, entity_type="patient", field_name=field)
    for field in TOKEN_TARGETS["sample"]:
        add(metadata.get(field), tokens.get("sample_token", ""), explicit_id=True, entity_type="sample", field_name=field)
    for field in TOKEN_TARGETS["report"]:
        add(metadata.get(field), tokens.get("report_token", ""), explicit_id=True, entity_type="report", field_name=field)

    add(metadata.get("patient_name"), "PATIENT_ANON")
    add(metadata.get("patient_birth_date"), "[REDACTED_BIRTH_DATE]")
    add(metadata.get("patient_birth_date_raw"), "[REDACTED_BIRTH_DATE]")
    add(metadata.get("patient_address"), "[REDACTED_ADDRESS]")

    add(metadata.get("prescriber"), "[REDACTED_STAFF]")
    add(metadata.get("validated_by"), "[REDACTED_STAFF]")
    add(metadata.get("edited_by"), "[REDACTED_STAFF]")
    add(metadata.get("printed_by"), "[REDACTED_STAFF]")

    add(metadata.get("phone"), "[REDACTED_CONTACT]")
    add(metadata.get("fax"), "[REDACTED_CONTACT]")

    return specs, originals


def anonymize_provenance(
    provenance: Dict[str, Any],
    specs: Sequence[ReplaceSpec],
    counters: Dict[str, int],
) -> Tuple[Dict[str, Any], int]:
    p = dict(provenance or {})
    replacements = 0

    for key in ("source_pdf", "extraction_json"):
        value = p.get(key)
        if isinstance(value, str):
            replaced, count = anonymize_text_safely(value, specs, counters)
            if count > 0:
                p[key] = replaced
                p[f"{key}_anonymized"] = replaced
                replacements += count

    return p, replacements


def anonymize_chunk(
    chunk: Dict[str, Any],
    mapping_store: MappingStore,
    counters: Dict[str, int],
) -> Tuple[Dict[str, Any], List[str]]:
    out = json.loads(json.dumps(chunk))
    metadata = dict(out.get("metadata") or {})
    provenance = dict(out.get("provenance") or {})

    doc_id = clean(out.get("doc_id"))

    tokens: Dict[str, str] = {}

    # Patient token (same token across patient_id / patient_id_raw / ip_patient if same raw value).
    for field in TOKEN_TARGETS["patient"]:
        raw = clean(metadata.get(field))
        if not raw:
            continue
        token, reused = mapping_store.get_or_create_token(
            entity_type="patient",
            field_name=field,
            original_value=raw,
            doc_id=doc_id,
            notes="patient identifier token",
        )
        tokens["patient_token"] = token
        counters["fields_tokenized_count"] += 1
        counters["tokens_reused" if reused else "tokens_created"] += 1

    for field in TOKEN_TARGETS["sample"]:
        raw = clean(metadata.get(field))
        if not raw:
            continue
        token, reused = mapping_store.get_or_create_token(
            entity_type="sample",
            field_name=field,
            original_value=raw,
            doc_id=doc_id,
            notes="sample identifier token",
        )
        tokens["sample_token"] = token
        counters["fields_tokenized_count"] += 1
        counters["tokens_reused" if reused else "tokens_created"] += 1

    for field in TOKEN_TARGETS["report"]:
        raw = clean(metadata.get(field))
        if not raw:
            continue
        token, reused = mapping_store.get_or_create_token(
            entity_type="report",
            field_name=field,
            original_value=raw,
            doc_id=doc_id,
            notes="report identifier token",
        )
        tokens["report_token"] = token
        counters["fields_tokenized_count"] += 1
        counters["tokens_reused" if reused else "tokens_created"] += 1

    if "patient_token" in tokens:
        metadata["patient_token"] = tokens["patient_token"]
    if "sample_token" in tokens:
        metadata["sample_token"] = tokens["sample_token"]
    if "report_token" in tokens:
        metadata["report_token"] = tokens["report_token"]

    # Remove original direct identifiers from metadata.
    for field in ("patient_id", "patient_id_raw", "ip_patient", "sample_id", "sample_id_raw", "report_id", "report_id_raw"):
        metadata.pop(field, None)

    # Redactions / neutralization.
    if clean(metadata.get("patient_name")):
        metadata["patient_name"] = "PATIENT_ANON"
        counters["fields_redacted_count"] += 1

    for field in (
        "patient_birth_date",
        "patient_birth_date_raw",
        "patient_address",
        "prescriber",
        "validated_by",
        "edited_by",
        "printed_by",
        "phone",
        "fax",
        "print_date",
    ):
        if clean(metadata.get(field)):
            counters["fields_redacted_count"] += 1
        metadata.pop(field, None)

    specs, sensitive_originals = build_replacement_specs(chunk.get("metadata") or {}, tokens)

    website_before = clean(metadata.get("website"))
    if website_before and maybe_sensitive_website(website_before, sensitive_originals):
        metadata["website"] = None
        counters["fields_redacted_count"] += 1

    embedding = clean(out.get("text_for_embedding"))
    keyword = clean(out.get("text_for_keyword"))

    embedding_replaced, c1 = anonymize_text_safely(embedding, specs, counters)
    keyword_replaced, c2 = anonymize_text_safely(keyword, specs, counters)
    embedding_replaced = sanitize_administrative_text(embedding_replaced)
    keyword_replaced = sanitize_administrative_text(keyword_replaced)
    counters["text_replacements_count"] += c1 + c2

    out["text_for_embedding"] = embedding_replaced
    out["text_for_keyword"] = keyword_replaced

    anon_prov, cp = anonymize_provenance(provenance, specs, counters)
    counters["text_replacements_count"] += cp
    out["provenance"] = anon_prov

    # Rebuild dedup key after anonymization.
    old_dedup = clean(metadata.get("dedup_key"))
    if old_dedup:
        metadata["source_dedup_key"] = old_dedup

    if out.get("chunk_type") in {"lab_result", "clinical_result"}:
        analyte_norm = clean(metadata.get("analyte_norm") or metadata.get("parameter_norm") or "unknown")
        value_raw = clean(metadata.get("value_raw")) or "na"
        unit = clean(metadata.get("unit"))
        if not unit:
            unit = "unit_missing"
        metadata["dedup_key"] = f"{analyte_norm}|{value_raw}|{unit}|{doc_id}"

    out["metadata"] = metadata
    out["schema_version"] = SCHEMA_VERSION
    out["content_hash"] = compute_content_hash(out)

    return out, sensitive_originals


def validate_anonymized(
    chunks: Sequence[Dict[str, Any]],
    sensitive_values_by_chunk: Dict[str, List[str]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], int]:
    errors: List[Dict[str, Any]] = []
    warnings: List[Dict[str, Any]] = []

    seen_ids: Set[str] = set()
    dup_ids: Set[str] = set()

    corrupted_patterns = [
        r"PAT_PAT_",
        r"SAMPLE_SAMPLE_",
        r"REPORT_REPORT_",
        r"PAT_[0-9]{6}PAT_",
        r"SAMPLE_[0-9]{6}SAMPLE_",
        r"REPORT_[0-9]{6}REPORT_",
        r"\d*PAT_[0-9]{6}\d*",
        r"PAT_[0-9]{6},\d+",
    ]
    corrupted_token_patterns_count = 0

    for chunk in chunks:
        chunk_id = clean(chunk.get("chunk_id"))
        doc_id = clean(chunk.get("doc_id"))

        missing = sorted([x for x in REQUIRED_CHUNK_FIELDS if x not in chunk])
        if missing:
            errors.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": f"missing_fields:{','.join(missing)}"})

        if not clean(chunk.get("text_for_embedding")):
            errors.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "empty_text_for_embedding"})

        if not clean(chunk.get("text_for_keyword")):
            errors.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "empty_text_for_keyword"})

        stored_hash = clean(chunk.get("content_hash"))
        if not stored_hash:
            errors.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "missing_content_hash"})
        else:
            expected_hash = compute_content_hash(chunk)
            if stored_hash != expected_hash:
                errors.append(
                    {
                        "chunk_id": chunk_id,
                        "doc_id": doc_id,
                        "issue": "content_hash_mismatch",
                    }
                )

        if chunk_id in seen_ids:
            dup_ids.add(chunk_id)
        seen_ids.add(chunk_id)

        metadata = chunk.get("metadata") or {}

        for field in FORBIDDEN_METADATA_FIELDS:
            if field in metadata:
                errors.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": f"forbidden_metadata_present:{field}"})

        if "patient_name" in metadata and clean(metadata.get("patient_name")) not in {"", "PATIENT_ANON"}:
            errors.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "patient_name_not_anonymized"})

        chunk_sensitive = sensitive_values_by_chunk.get(chunk_id, [])
        text_blob = f"{clean(chunk.get('text_for_embedding'))} {clean(chunk.get('text_for_keyword'))}"
        for value in sorted(set(x for x in chunk_sensitive if clean(x)), key=len, reverse=True):
            if len(value) < 4:
                continue
            if value in text_blob:
                errors.append({
                    "chunk_id": chunk_id,
                    "doc_id": doc_id,
                    "issue": "sensitive_value_found_in_text",
                    "value_preview": value[:40],
                })
                break

        for pattern in corrupted_patterns:
            if re.search(pattern, text_blob):
                corrupted_token_patterns_count += 1
                errors.append({
                    "chunk_id": chunk_id,
                    "doc_id": doc_id,
                    "issue": "corrupted_token_pattern_detected",
                    "pattern": pattern,
                })
                break

        for pattern in BLOCKING_ADMIN_PATTERNS:
            if re.search(pattern, text_blob, flags=re.IGNORECASE):
                errors.append({
                    "chunk_id": chunk_id,
                    "doc_id": doc_id,
                    "issue": "administrative_label_found_in_indexable_text",
                    "pattern": pattern,
                })
                break

        if clean(metadata.get("validation_status")) == "unknown":
            warnings.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "validation_status_unknown"})

        if clean(metadata.get("unit_quality_status")) == "unit_missing":
            warnings.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "unit_quality_status_unit_missing"})

        if clean(metadata.get("reference_quality_status")) == "possibly_truncated":
            warnings.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "reference_quality_status_possibly_truncated"})

        if clean(metadata.get("age_consistency_status")) == "inconsistent_with_birth_date":
            warnings.append({"chunk_id": chunk_id, "doc_id": doc_id, "issue": "age_consistency_status_inconsistent_with_birth_date"})

    if dup_ids:
        for cid in sorted(dup_ids):
            errors.append({"chunk_id": cid, "doc_id": "", "issue": "duplicate_chunk_id"})

    chunk_ids = {clean(c.get("chunk_id")) for c in chunks}
    for chunk in chunks:
        parent = clean(chunk.get("parent_chunk_id"))
        if parent and parent not in chunk_ids:
            errors.append({
                "chunk_id": clean(chunk.get("chunk_id")),
                "doc_id": clean(chunk.get("doc_id")),
                "issue": "broken_parent_chunk_id",
                "parent_chunk_id": parent,
            })

    return errors, warnings, corrupted_token_patterns_count


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Anonymize raw chunk JSONL before indexing.")
    parser.add_argument("--input", type=Path, required=True, help="Input raw JSONL path")
    parser.add_argument("--output", type=Path, required=True, help="Output anonymized JSONL path")
    parser.add_argument("--mapping-xlsx", type=Path, required=True, help="Private reversible mapping XLSX path")
    parser.add_argument("--report", type=Path, required=True, help="Output anonymization report path")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not args.input.exists():
        print(f"ERROR - input file does not exist: {args.input}", file=sys.stderr)
        return 1

    print(f"INFO - loading raw chunks: {args.input}")
    try:
        raw_chunks = load_jsonl(args.input)
    except Exception as exc:
        report = {
            "schema_version": SCHEMA_VERSION,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "input_path": str(args.input),
            "output_path": str(args.output),
            "mapping_path": str(args.mapping_xlsx),
            "total_chunks": 0,
            "chunks_by_type": {},
            "documents_count": 0,
            "tokens_created": 0,
            "tokens_reused": 0,
            "fields_tokenized_count": 0,
            "fields_redacted_count": 0,
            "text_replacements_count": 0,
            "skipped_short_numeric_replacements_count": 0,
            "corrupted_token_patterns_count": 0,
            "validation_errors": [{"issue": f"json_invalid:{exc}"}],
            "validation_warnings": [],
            "readiness_status": "blocked",
        }
        write_json(args.report, report)
        print(f"ERROR - invalid input JSONL: {exc}", file=sys.stderr)
        return 1

    mapping_store, loaded_mapping_path, load_warnings = read_mapping(args.mapping_xlsx)
    if load_warnings:
        for w in load_warnings:
            print(f"WARNING - {w}")

    counters: Dict[str, int] = defaultdict(int)
    anonymized_chunks: List[Dict[str, Any]] = []
    sensitive_values_by_chunk: Dict[str, List[str]] = {}

    for chunk in raw_chunks:
        anonymized, sensitive_values = anonymize_chunk(chunk, mapping_store, counters)
        cid = clean(anonymized.get("chunk_id"))
        sensitive_values_by_chunk[cid] = sensitive_values
        anonymized_chunks.append(anonymized)

    print(f"INFO - saving mapping: {args.mapping_xlsx}")
    mapping_path_written, save_warnings = save_mapping(args.mapping_xlsx, mapping_store)
    if save_warnings:
        for w in save_warnings:
            print(f"WARNING - {w}")

    print("INFO - validating anonymized chunks")
    validation_errors, validation_warnings, corrupted_token_patterns_count = validate_anonymized(
        anonymized_chunks,
        sensitive_values_by_chunk,
    )

    readiness_status = "ready_for_indexing" if not validation_errors else "blocked"

    chunks_by_type = Counter(clean(c.get("chunk_type")) for c in anonymized_chunks)
    docs = {clean(c.get("doc_id")) for c in anonymized_chunks if clean(c.get("doc_id"))}

    report = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "input_path": str(args.input),
        "output_path": str(args.output),
        "mapping_path": mapping_path_written,
        "mapping_loaded_from": loaded_mapping_path,
        "total_chunks": len(anonymized_chunks),
        "chunks_by_type": dict(chunks_by_type),
        "documents_count": len(docs),
        "tokens_created": counters["tokens_created"],
        "tokens_reused": counters["tokens_reused"],
        "fields_tokenized_count": counters["fields_tokenized_count"],
        "fields_redacted_count": counters["fields_redacted_count"],
        "text_replacements_count": counters["text_replacements_count"],
        "skipped_short_numeric_replacements_count": counters["skipped_short_numeric_replacements_count"],
        "corrupted_token_patterns_count": corrupted_token_patterns_count,
        "validation_errors": validation_errors,
        "validation_warnings": validation_warnings,
        "readiness_status": readiness_status,
        "notes": [
            "Raw chunks are never modified by this script.",
            "Only anonymized chunks should be indexed.",
            "Mapping is private and must never be committed.",
        ],
    }

    # Always write artifacts. Indexing decision must rely on readiness_status.
    write_jsonl(args.output, anonymized_chunks)
    write_json(args.report, report)

    print(f"INFO - chunks written: {len(anonymized_chunks)}")
    print(f"INFO - anonymized output: {args.output}")
    print(f"INFO - report: {args.report}")
    print(f"INFO - readiness: {readiness_status}")

    if readiness_status == "blocked":
        print(f"ERROR - blocking validation errors: {len(validation_errors)}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
