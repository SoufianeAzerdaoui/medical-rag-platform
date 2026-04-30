from __future__ import annotations

import csv
from dataclasses import replace
from pathlib import Path
from typing import Any

try:
    from .config import MAPPING_CSV, MAPPING_XLSX
    from .models import RetrievalFilters
except ImportError:
    from config import MAPPING_CSV, MAPPING_XLSX
    from models import RetrievalFilters


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _token_guess(prefix: str, raw_value: str) -> str | None:
    raw = _clean(raw_value)
    if not raw:
        return None
    if raw.upper().startswith(prefix + "_"):
        return raw.upper()
    if raw.isdigit() and 0 <= int(raw) <= 999999:
        return f"{prefix}_{int(raw):06d}"
    return None


class MappingResolver:
    """
    Resolve raw IDs (patient_id/sample_id/report_id) to anonymized tokens when mapping is available.
    """

    def __init__(self, csv_path: Path = MAPPING_CSV, xlsx_path: Path = MAPPING_XLSX) -> None:
        self.csv_path = csv_path
        self.xlsx_path = xlsx_path
        self.by_entity_original: dict[tuple[str, str], str] = {}
        self._load()

    def _load(self) -> None:
        if self.csv_path.exists():
            self._load_csv(self.csv_path)
            return
        if self.xlsx_path.exists():
            self._load_xlsx(self.xlsx_path)

    def _load_csv(self, path: Path) -> None:
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                entity = _clean(row.get("entity_type")).lower()
                original = _clean(row.get("original_value"))
                anon = _clean(row.get("anonymized_value"))
                if entity and original and anon:
                    self.by_entity_original[(entity, original)] = anon

    def _load_xlsx(self, path: Path) -> None:
        try:
            import openpyxl  # type: ignore
        except Exception:
            return
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [(_clean(c) or "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
        col_idx = {name: i for i, name in enumerate(header)}
        for vals in ws.iter_rows(min_row=2, values_only=True):
            entity = _clean(vals[col_idx.get("entity_type", -1)] if col_idx.get("entity_type") is not None else "").lower()
            original = _clean(vals[col_idx.get("original_value", -1)] if col_idx.get("original_value") is not None else "")
            anon = _clean(vals[col_idx.get("anonymized_value", -1)] if col_idx.get("anonymized_value") is not None else "")
            if entity and original and anon:
                self.by_entity_original[(entity, original)] = anon

    def resolve(self, entity_type: str, original_value: str | None) -> str | None:
        if not original_value:
            return None
        key = (entity_type.lower(), _clean(original_value))
        return self.by_entity_original.get(key)


def resolve_filters(filters: RetrievalFilters, resolver: MappingResolver | None = None) -> RetrievalFilters:
    f = replace(filters)
    resolver = resolver or MappingResolver()

    if f.patient_id:
        f.patient_token = resolver.resolve("patient", f.patient_id) or _token_guess("PAT", f.patient_id)

    if f.sample_id:
        # sample_id can map either as sample token or (legacy) report token values in some workflows
        f.sample_token = resolver.resolve("sample", f.sample_id) or _token_guess("SAMPLE", f.sample_id)
        if f.report_token is None:
            f.report_token = resolver.resolve("report", f.sample_id) or _token_guess("REPORT", f.sample_id)

    return f


def build_sql_filter_clauses(filters: RetrievalFilters) -> tuple[list[str], list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []

    def add(clause: str, *values: Any) -> None:
        clauses.append(clause)
        params.extend(values)

    if filters.document_type:
        add("m.document_type = ?", filters.document_type)
    if filters.doc_id:
        add("c.doc_id = ?", filters.doc_id)
    if filters.sample_type:
        add("m.sample_type = ?", filters.sample_type)
    if filters.chunk_type:
        add("c.chunk_type = ?", filters.chunk_type)
    if filters.source_pdf:
        add("COALESCE(m.source_pdf, o.source_pdf) = ?", filters.source_pdf)

    if filters.request_date:
        add("substr(m.request_date, 1, 10) = ?", filters.request_date)
    if filters.request_date_from:
        add("substr(m.request_date, 1, 10) >= ?", filters.request_date_from)
    if filters.request_date_to:
        add("substr(m.request_date, 1, 10) <= ?", filters.request_date_to)

    if filters.patient_token:
        add("m.patient_token = ?", filters.patient_token)
    if filters.sample_token:
        add("m.sample_token = ?", filters.sample_token)
    if filters.report_token:
        add("m.report_token = ?", filters.report_token)

    return clauses, params


def row_matches_filters(row: dict[str, Any], filters: RetrievalFilters) -> bool:
    def getv(*keys: str) -> str:
        for key in keys:
            if key in row and row[key] is not None:
                return str(row[key])
        return ""

    if filters.document_type and getv("document_type") != filters.document_type:
        return False
    if filters.doc_id and getv("doc_id") != filters.doc_id:
        return False
    if filters.sample_type and getv("sample_type") != filters.sample_type:
        return False
    if filters.chunk_type and getv("chunk_type") != filters.chunk_type:
        return False
    if filters.source_pdf and getv("source_pdf") != filters.source_pdf:
        return False

    req_date = getv("request_date")
    req_day = req_date[:10] if req_date else ""
    if filters.request_date and req_day != filters.request_date:
        return False
    if filters.request_date_from and req_day and req_day < filters.request_date_from:
        return False
    if filters.request_date_to and req_day and req_day > filters.request_date_to:
        return False

    if filters.patient_token and getv("patient_token") != filters.patient_token:
        return False
    if filters.sample_token and getv("sample_token") != filters.sample_token:
        return False
    if filters.report_token and getv("report_token") != filters.report_token:
        return False

    return True
